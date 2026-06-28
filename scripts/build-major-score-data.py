#!/usr/bin/env python3

from __future__ import annotations

import argparse
import gzip
import json
import math
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


SOURCE_DIR = Path("各专业分数线")
OUTPUT_DIR = Path("data/major-scores")

PROVINCE_NAMES = {
    "北京": ("北京市", "beijing"),
    "天津": ("天津市", "tianjin"),
    "上海": ("上海市", "shanghai"),
    "重庆": ("重庆市", "chongqing"),
    "河北": ("河北省", "hebei"),
    "山西": ("山西省", "shanxi"),
    "辽宁": ("辽宁省", "liaoning"),
    "吉林": ("吉林省", "jilin"),
    "黑龙江": ("黑龙江省", "heilongjiang"),
    "江苏": ("江苏省", "jiangsu"),
    "浙江": ("浙江省", "zhejiang"),
    "安徽": ("安徽省", "anhui"),
    "福建": ("福建省", "fujian"),
    "江西": ("江西省", "jiangxi"),
    "山东": ("山东省", "shandong"),
    "河南": ("河南省", "henan"),
    "湖北": ("湖北省", "hubei"),
    "湖南": ("湖南省", "hunan"),
    "广东": ("广东省", "guangdong"),
    "海南": ("海南省", "hainan"),
    "四川": ("四川省", "sichuan"),
    "贵州": ("贵州省", "guizhou"),
    "云南": ("云南省", "yunnan"),
    "陕西": ("陕西省", "shaanxi"),
    "甘肃": ("甘肃省", "gansu"),
    "青海": ("青海省", "qinghai"),
    "内蒙古": ("内蒙古自治区", "neimenggu"),
    "广西": ("广西壮族自治区", "guangxi"),
    "西藏": ("西藏自治区", "xizang"),
    "宁夏": ("宁夏回族自治区", "ningxia"),
    "新疆": ("新疆维吾尔自治区", "xinjiang"),
}

SUBJECT_CODES = {"物理": 0, "历史": 1, "综合": 2}
LEVEL_CODES = {"本科": 0, "专科": 1}


def main() -> None:
    parser = argparse.ArgumentParser(description="Build compact major admission score data from provincial xlsx files.")
    parser.add_argument("--source-dir", default=str(SOURCE_DIR))
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    parser.add_argument("--province", action="append", help="Only build matching short province name, e.g. 河南")
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale_path in list(output_dir.glob("*.json")) + list(output_dir.glob("*.json.gz")):
        stale_path.unlink()

    wanted = set(args.province or [])
    manifest = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": "各专业分数线/22-25年全国高校在各省的专业录取分数",
        "version": 1,
        "subjectCodes": {value: key for key, value in SUBJECT_CODES.items()},
        "levelCodes": {value: key for key, value in LEVEL_CODES.items()},
        "provinces": {},
    }

    for workbook_path in sorted(source_dir.glob("*.xlsx")):
        short_name = extract_short_province(workbook_path.name)
        if not short_name or (wanted and short_name not in wanted):
            continue
        province_name, slug = PROVINCE_NAMES.get(short_name, (f"{short_name}省", short_name))
        print(f"Building {province_name} from {workbook_path.name}")
        province_data = build_province_data(workbook_path, province_name)
        output_name = f"{slug}.json.gz"
        output_path = output_dir / output_name
        encoded = json.dumps(province_data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        output_path.write_bytes(gzip.compress(encoded, compresslevel=9))
        manifest["provinces"][province_name] = {
            "file": output_name,
            "encoding": "gzip",
            "shortName": short_name,
            "rawRows": province_data["meta"]["rawRows"],
            "keptRows": province_data["meta"]["keptRows"],
            "records": province_data["meta"]["records"],
            "years": province_data["meta"]["years"],
            "subjects": province_data["meta"]["subjects"],
            "sourceFile": workbook_path.name,
        }

    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote manifest for {len(manifest['provinces'])} provinces to {output_dir / 'manifest.json'}")


def build_province_data(path: Path, province_name: str) -> dict:
    dictionaries = {
        "schools": Dictionary(),
        "majors": Dictionary(),
        "batches": Dictionary(),
        "groups": Dictionary(),
        "requirements": Dictionary(),
        "locations": Dictionary(),
        "ownerships": Dictionary(),
    }
    raw_rows = 0
    kept_rows = 0
    skipped_rows = defaultdict(int)
    subjects = defaultdict(int)
    years = defaultdict(int)
    grouped: dict[tuple, dict[int, dict]] = defaultdict(dict)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [normalize_header(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: pos for pos, name in enumerate(header)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_rows += 1
        item = {key: row[pos] if pos < len(row) else None for key, pos in index.items()}
        year = number_value(item.get("year"))
        score = number_value(item.get("score"))
        subject = normalize_subject(item.get("subject"))
        level = infer_level(item.get("batch"))
        school = clean_text(item.get("school"))
        major = clean_text(item.get("major"))

        if not year or year < 2022 or year > 2025:
            skipped_rows["year"] += 1
            continue
        if subject not in SUBJECT_CODES:
            skipped_rows["subject"] += 1
            continue
        if not school or not major:
            skipped_rows["name"] += 1
            continue
        if not score or score < 1 or score > 900:
            skipped_rows["score"] += 1
            continue

        kept_rows += 1
        subjects[subject] += 1
        years[str(year)] += 1

        batch = clean_text(item.get("batch"))
        group = clean_text(item.get("group"))
        requirement = clean_text(item.get("requirement"))
        location = normalize_location(item.get("location"))
        ownership = clean_text(item.get("ownership")) or "未标注"
        plan = number_value(item.get("plan"))
        rank = number_value(item.get("rank"))
        is_985 = yes_no(item.get("is985"))
        is_211 = yes_no(item.get("is211"))

        key = (
            dictionaries["schools"].id(school),
            dictionaries["majors"].id(major),
            SUBJECT_CODES[subject],
            dictionaries["batches"].id(batch),
            dictionaries["groups"].id(group),
            dictionaries["requirements"].id(requirement),
            LEVEL_CODES[level],
            dictionaries["locations"].id(location),
            dictionaries["ownerships"].id(ownership),
            (1 if is_985 else 0) | (2 if is_211 else 0),
        )

        current = grouped[key].get(year)
        if current is None or score < current["score"]:
            grouped[key][year] = {"score": score, "rank": rank, "plan": plan}
        elif current is not None and plan:
            current["plan"] = (current.get("plan") or 0) + plan

    wb.close()

    records = []
    for key, history_map in grouped.items():
        history = []
        for year, values in sorted(history_map.items(), reverse=True):
            history.extend([year, values["score"], values.get("rank") or 0, values.get("plan") or 0])
        records.append([*key, history])

    records.sort(key=lambda record: (
        record[0],
        record[4],
        record[2],
        -record[-1][0],
        record[-1][1],
    ))

    return {
        "meta": {
            "province": province_name,
            "rawRows": raw_rows,
            "keptRows": kept_rows,
            "records": len(records),
            "years": sorted(years.keys()),
            "subjects": dict(sorted(subjects.items())),
            "skippedRows": dict(sorted(skipped_rows.items())),
        },
        "dict": {name: dictionary.values for name, dictionary in dictionaries.items()},
        "records": records,
    }


class Dictionary:
    def __init__(self) -> None:
        self.values: list[str] = []
        self._ids: dict[str, int] = {}

    def id(self, value: str) -> int:
        value = value or ""
        existing = self._ids.get(value)
        if existing is not None:
            return existing
        new_id = len(self.values)
        self.values.append(value)
        self._ids[value] = new_id
        return new_id


def extract_short_province(file_name: str) -> str:
    match = re.search(r"在(.+?)的专业录取分数", file_name)
    return match.group(1) if match else ""


def normalize_header(value) -> str:
    key = clean_text(value).replace(" ", "")
    mapping = {
        "年份": "year",
        "院校名称": "school",
        "院校代码": "schoolCode",
        "科类": "subject",
        "批次": "batch",
        "专业": "major",
        "专业代码": "majorCode",
        "所属专业组": "group",
        "专业备注": "remark",
        "选科要求": "requirement",
        "录取人数": "plan",
        "最低分数": "score",
        "最低位次": "rank",
        "学校所在": "location",
        "学校性质": "ownership",
        "是否985": "is985",
        "是否211": "is211",
    }
    return mapping.get(key, key)


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return re.sub(r"\s+", " ", text)


def number_value(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)) and not math.isnan(value):
        return int(round(value))
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    return int(round(float(match.group(0)))) if match else 0


def normalize_subject(value) -> str:
    text = clean_text(value)
    if "艺术" in text or "体育" in text or "三校" in text or "蒙授" in text:
        return ""
    if "物理" in text or "理科" in text or text == "理":
        return "物理"
    if "历史" in text or "文科" in text or text == "文":
        return "历史"
    if "综合" in text:
        return "综合"
    return text


def infer_level(batch) -> str:
    text = clean_text(batch)
    return "专科" if "专科" in text or "高职" in text else "本科"


def normalize_location(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if text in PROVINCE_NAMES:
        return PROVINCE_NAMES[text][0]
    return text


def yes_no(value) -> bool:
    return clean_text(value) in {"是", "1", "true", "True", "TRUE", "Y", "yes"}


if __name__ == "__main__":
    main()
